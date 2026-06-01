import io
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from app import models
from app.database import get_db
from app.dependencies import get_current_user
from app.services.venta_service import obtener_ventas_empresa

router = APIRouter(prefix="/reportes", tags=["Reportes"])

_HEADER_FILL = PatternFill("solid", fgColor="A78BFA")   # neiva-purple
_HEADER_FONT = Font(bold=True, color="FFFFFF")


@router.get("/ventas/excel/{empresa_id}")
def exportar_ventas_excel(
    empresa_id: UUID,
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user),
):
    if current_user.empresa_id != empresa_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Acceso no autorizado",
        )

    ventas = obtener_ventas_empresa(empresa_id, db)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"

    # Cabeceras
    columnas = [
        "ID Venta", "Fecha", "Producto",
        "Cantidad", "Precio Unitario", "Subtotal", "Total Venta",
    ]
    ws.append(columnas)

    for celda in ws[1]:
        celda.font = _HEADER_FONT
        celda.fill = _HEADER_FILL
        celda.alignment = Alignment(horizontal="center")

    # Filas de datos
    for venta in ventas:
        for detalle in venta.detalles:
            ws.append([
                str(venta.id),
                venta.fecha_venta.strftime("%Y-%m-%d %H:%M"),
                detalle.producto.nombre if detalle.producto else "",
                detalle.cantidad,
                float(detalle.precio_unitario),
                float(detalle.subtotal),
                float(venta.total),
            ])

    # Ajustar ancho de columnas automáticamente
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": f"attachment; filename=ventas_{empresa_id}.xlsx"
        },
    )


@router.get("/financieros/{empresa_id}")
def obtener_resumen_financiero(
    empresa_id: UUID,
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user),
):
    if current_user.empresa_id != empresa_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Acceso no autorizado",
        )

    from sqlalchemy import func

    # 1. Inversion en inventario activo (stock * costo)
    inversion_activa = db.query(
        func.sum(models.Producto.cantidad_actual * models.Producto.precio_costo)
    ).filter(
        models.Producto.empresa_id == empresa_id,
        models.Producto.is_active.is_(True),
    ).scalar() or 0

    # 2. Ingresos totales por ventas
    ingresos_totales = db.query(
        func.sum(models.Venta.total)
    ).filter(
        models.Venta.empresa_id == empresa_id,
    ).scalar() or 0

    # 3. Costo de ventas (COGS acumulado)
    cogs_total = db.query(
        func.sum(models.DetalleVenta.cantidad * models.Producto.precio_costo)
    ).join(
        models.Producto,
        models.DetalleVenta.producto_id == models.Producto.id,
    ).filter(
        models.Producto.empresa_id == empresa_id,
    ).scalar() or 0

    # 4. Ganancia neta realizada (ingresos - costo de lo vendido)
    ganancia_neta = float(ingresos_totales) - float(cogs_total)

    # 5. Costo de reposicion para productos con stock bajo (<= 5) para llevarlos a stock ideal de 15
    productos_bajo_stock = db.query(models.Producto).filter(
        models.Producto.empresa_id == empresa_id,
        models.Producto.is_active.is_(True),
        models.Producto.cantidad_actual <= 5,
    ).all()

    costo_reposicion_bajo_stock = sum(
        max(0.0, 15.0 - float(p.cantidad_actual)) * float(p.precio_costo)
        for p in productos_bajo_stock
    )

    return {
        "inversion_activa": float(inversion_activa),
        "ingresos_totales": float(ingresos_totales),
        "cogs_total": float(cogs_total),
        "ganancia_neta": float(ganancia_neta),
        "costo_reposicion_bajo_stock": float(costo_reposicion_bajo_stock),
        "costo_reposicion_total_sugerido": float(cogs_total) + float(costo_reposicion_bajo_stock),
    }
